"""Functions for extracting data from AFIP."""
from selenium.webdriver.common.by import By
import openpyxl
import os

def extract_table(driver):
    """Extract data from the results table."""
    print("📝 EXTRAYENDO DATOS")
    try:
        tabla = driver.find_element(By.ID, "tablaDataTables")
        tbody = tabla.find_element(By.TAG_NAME, "tbody")
        filas = tbody.find_elements(By.TAG_NAME, "tr")
        
        # Extract data from the table
        datos = [[celda.text for celda in fila.find_elements(By.TAG_NAME, "td")] for fila in filas]
        
        # Remove empty rows
        datos = [fila for fila in datos if fila]
        return datos
    except Exception as e:
        print(f"Error al extraer datos de la tabla: {e}")
        return []

def extract_table_data(driver):
    """Extract data from the results table."""
    print("📝 EXTRAYENDO DATOS")
    try:
        tabla = driver.find_element(By.CLASS_NAME, "jig_table")
        tbody = tabla.find_element(By.TAG_NAME, "tbody")
        filas = tbody.find_elements(By.TAG_NAME, "tr")
        
        # Extract data from the table
        datos = [[celda.text for celda in fila.find_elements(By.TAG_NAME, "td")] for fila in filas]
        
        # Remove empty rows
        datos = [fila for fila in datos if fila]
        return datos
    except Exception as e:
        print(f"Error al extraer datos de la tabla: {e}")
        return []

def save_to_excel(datos, filename):
    """Save data to an Excel file."""
    print("📝 GENERANDO EXCEL")
    try:
        # Verificar si el archivo existe
        if os.path.exists(filename):
            libro = openpyxl.load_workbook(filename)  # Cargar archivo existente
            hoja = libro.active
        else:
            libro = openpyxl.Workbook()  # Crear nuevo archivo
            hoja = libro.active
        
        # Determinar la última fila con datos
        ultima_fila = hoja.max_row

        # Agregar datos al final del archivo
        for fila_idx, fila in enumerate(datos, start=ultima_fila + 1):
            for col_idx, valor in enumerate(fila, start=1):
                hoja.cell(row=fila_idx, column=col_idx, value=valor)
        
        # Guardar cambios
        libro.save(filename)
        print("🎉 FILAS AGREGADAS AL EXCEL!!")
        return True
    except Exception as e:
        print(f"Error al agregar filas al archivo Excel: {e}")
        return False